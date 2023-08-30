from tkinter import *
from tkinter import filedialog
from tkinter import messagebox

import openpyxl
import random

class programa_spe:

    def __init__(self):
        self.root = Tk()
        self.root.title("Programa SPE")
        self.root.iconbitmap("logo.ico")

        self.bg = "#0c75b7"
        self.root.config(bg=self.bg)
        
        self.font = "Arial Black"
        self.font_text = "Aerial"

        ancho = 1000
        alto = 600

        #----------------colocar al centro de la pantalla-------------------#
        self.centrar_interfaz(ancho, alto)
        
        self.root.minsize(ancho, alto)

        #----------------frames-------------------#
        self.frame_inicio = Frame(self.root)
        self.frame_cargar = Frame(self.root)
        self.frame_datos = Frame(self.root)
        self.frame_preguntas = Frame(self.root)
        self.frame_puntuacion = Frame(self.root)

        #-------------create widgets--------------#
        self.widgets_inicio()
        self.widgets_cargar()
        self.widgets_datos()
        self.widgets_preguntas()
        self.widgets_puntuacion()

        self.frame_inicio.place(anchor="c", relx=.5, rely=.5)
        
        #-------------si presiona Enter llama a la funcion comenzar_juego--------------#
        self.root.bind("<Key-Return>", lambda _: self.comenzar_juego())
        #-------------si presiona Escape llama a la funcion salir_programa-------------#
        self.root.bind("<Key-Escape>", lambda _: self.salir_programa())

    
    #----------------centrar interfaz-------------------#
    def centrar_interfaz(self, ancho, alto):

        #------------------dimensiones de la pantalla-------------------#
        ancho_pantalla = self.root.winfo_screenwidth()
        alto_pantalla = self.root.winfo_screenheight()

        #----------------------centro de la pantalla---------------------#
        x = int((ancho_pantalla/2) - (ancho/2))
        y = int((alto_pantalla/2) - (alto/2))

        #---------colocar interfaz en el centro de la pantalla-----------#
        self.root.geometry(f'{ancho}x{alto}+{x}+{y}')
        

    #------------------------create widgets----------------------------#

    def widgets_inicio(self):
        #-------------backgraund and geometry--------------#
        self.frame_inicio.config(bg="#0c75b7")

        self.logo_petro = PhotoImage(file="logo-petrobowl.png")
        label_petro = Label(self.frame_inicio, image=self.logo_petro, background=self.bg)
        label_petro.grid(row=0, column=0, padx=30, columnspan=3)

        self.logo_spe = PhotoImage(file="logo-spe.png")
        label_spe = Label(self.frame_inicio, image=self.logo_spe, background=self.bg)
        label_spe.grid(row=0, column=0, padx=30, rowspan=5)

        self.logo_udo = PhotoImage(file="logo-udo.png")
        label_udo = Label(self.frame_inicio, image=self.logo_udo, background=self.bg)
        label_udo.grid(row=0, column=2, padx=30, rowspan=5)

        #---------------------labels-----------------------#
        label_bienvenidos = Label(self.frame_inicio, text="Bienvenidos", fg="#ffffff", background=self.bg, font=("Monotype Corsiva", 55))
        label_bienvenidos.grid(row=1, column=1, padx=10, pady=60)

        #---------------------buttons-----------------------#
        self.comenzar = Button(self.frame_inicio, text="Comenzar", font=(self.font, 12), command=self.comenzar_juego)
        self.comenzar.grid(row=2, column=1, padx=10, pady=10)

        acerca = Button(self.frame_inicio, text="Acerca de", font=(self.font, 12), command=self.acerca_de)
        acerca.grid(row=3, column=1, padx=10, pady=10)

        salir = Button(self.frame_inicio, text="Salir", font=(self.font, 12), command=self.salir_programa)
        salir.grid(row=4, column=1, padx=10, pady=10)

    
    def widgets_cargar(self):
        #-------------backgraund and geometry--------------#
        self.frame_cargar.config(bg=self.bg)

        #----------------------labels----------------------#
        label_cargar = Label(self.frame_cargar, text="Cargar preguntas", fg="#ffffff", background=self.bg, font=("Monotype Corsiva", 45))
        label_cargar.grid(row=0, column=0, pady=10, columnspan=6)

        #----------------------entry-----------------------#
        self.direccion = StringVar()
        self.direccion.set("")

        self.entry_direccion = Entry(self.frame_cargar, width=50, font=(self.font_text, 12), state="disabled", textvariable=self.direccion)
        self.entry_direccion.grid(row=1, column=0, padx=10, pady=10)

        #----------------------buttons---------------------#
        cargar = Button(self.frame_cargar, text="Cargar", font=(self.font, 10), command=self.cargar_preguntas)
        cargar.grid(row=1, column=2, padx=10, pady=10, columnspan=5)

        self.iniciar = Button(self.frame_cargar, text="Iniciar", font=(self.font, 10), command=self.cargar_participantes, state=DISABLED)
        self.iniciar.grid(row=4, column=0, pady=10, columnspan=5)

        volver = Button(self.frame_cargar, text="Volver", font=(self.font, 10), command=self.volver_inicio)
        volver.grid(row=5, column=0, pady=10, columnspan=5)

    
    def widgets_datos(self):

        self.frame_datos.config(bg=self.bg)

        label_participantes = Label(self.frame_datos, text="Datos de la ronda", fg="#ffffff", background=self.bg, font=("Monotype Corsiva", 55))
        label_participantes.grid(row=0, column=0, pady=5)

        self.p1 = StringVar(value="")
        self.p2 = StringVar(value="")

        label_participante_1 = Label(self.frame_datos, text="Participante 1:", fg="#ffffff", background=self.bg, font=(self.font_text, 15))
        label_participante_1.grid(row=1, column=0)

        entry_participante_1 = Entry(self.frame_datos, width=20, textvariable=self.p1, font=("Arial", 15))
        entry_participante_1.grid(row=2, column=0, pady=10)

        label_participante_1 = Label(self.frame_datos, text="Participante 2:", fg="#ffffff", background=self.bg, font=(self.font_text, 15))
        label_participante_1.grid(row=3, column=0)

        entry_participante_2 = Entry(self.frame_datos, width=20, textvariable=self.p2, font=(self.font_text, 15))
        entry_participante_2.grid(row=4, column=0, pady=10)


        self.c_preguntas = StringVar()

        label_cantidad = Label(self.frame_datos, text="Cantidad de preguntas", fg="#ffffff", background=self.bg, font=("Monotype Corsiva", 20))
        label_cantidad.grid(row=5, column=0, pady=10, columnspan=5)

        self.entry_cantidad = Entry(self.frame_datos, width=3, font=(self.font, 12), justify="center", textvariable=self.c_preguntas)
        self.entry_cantidad.grid(row=6, column=0, pady=10, columnspan=5)


        self.comenzar_ronda = Button(self.frame_datos, text="Comenzar ronda", font=(self.font, 12), command=self.iniciar_ronda)
        self.comenzar_ronda.grid(row=7, column=0, pady=10)

        volver = Button(self.frame_datos, text="Volver", font=(self.font, 12), command=self.volver_cargar)
        volver.grid(row=8, column=0, pady=10)


    def widgets_preguntas(self):

        #-------------backgraund and geometry--------------#
        self.frame_preguntas.config(bg=self.bg)

        #---------------------labels-----------------------#
        label_p1 = Label(self.frame_preguntas, textvariable=self.p1, fg="#ffffff", background=self.bg, font=(self.font, 30))
        label_p1.grid(row=0, column=0, padx=5, pady=10, sticky="e")

        label_p2 = Label(self.frame_preguntas, textvariable=self.p2, fg="#ffffff", background=self.bg, font=(self.font, 30))
        label_p2.grid(row=0, column=3, padx=5, pady=10, sticky="w")
        
        self.tiempo = IntVar(value="")

        self.label_tiempo = Label(self.frame_preguntas, background=self.bg, fg="#ffffff", font=(self.font, 30), textvariable=self.tiempo)
        self.label_tiempo.grid(row=0, column=1, columnspan=2)

        self.label_pregunta = Label(self.frame_preguntas, text="Pregunta:", fg="#ffffff", background=self.bg, font=(self.font_text, 20))
        self.label_pregunta.grid(row=1, column=0, padx=5, sticky="e")
        self.label_respuesta = Label(self.frame_preguntas, text="Respuesta:", fg="#ffffff", background=self.bg, font=(self.font_text, 20))
        self.label_respuesta.grid(row=2, column=0, padx=5, sticky="e")

        #---------------------entrys-----------------------#
        self.pregunta = StringVar()
        self.pregunta.set("")

        self.text_pregunta = Text(self.frame_preguntas, width=50, height=5, font=(self.font_text, 12), state="disabled")
        self.text_pregunta.grid(row=1, column=1, pady=10)
        
        scroll_pregunta = Scrollbar(self.frame_preguntas, command=self.text_pregunta.yview)
        scroll_pregunta.grid(row=1, column=2, sticky="nsew", pady=10)
        self.text_pregunta.config(yscrollcommand=scroll_pregunta.set)

        self.text_respuesta = Text(self.frame_preguntas, width=50, height=5, font=(self.font_text, 12), state="disabled")
        self.text_respuesta.grid(row=2, column=1, pady=20)
        
        scroll_respuesta = Scrollbar(self.frame_preguntas, command=self.text_pregunta.yview)
        scroll_respuesta.grid(row=2, column=2, sticky="nsew", pady=20)
        self.text_respuesta.config(yscrollcommand=scroll_respuesta.set)

        #---------------------buttons-----------------------#
        self.responder_1 = Button(self.frame_preguntas, text="Responder 1", font=(self.font, 10), command= lambda: self.temporizador(0))
        self.responder_1.grid(row=3, column=1, pady=10, sticky="w")

        self.responder_2 = Button(self.frame_preguntas, text="Responder 2", font=(self.font, 10), command= lambda: self.temporizador(1))
        self.responder_2.grid(row=3, column=1, pady=10, sticky="e")

        self.correcto = Button(self.frame_preguntas, text="Correcto", font=(self.font, 10), command=self.r_correcta)
        self.correcto.grid(row=4, column=1)

        self.incorrecto = Button(self.frame_preguntas, text="Incorrecto", font=(self.font, 10), command=self.r_incorrecta)
        self.incorrecto.grid(row=5, column=1, pady=10)

        omitir = Button(self.frame_preguntas, text="Omitir", font=(self.font, 10), command=self.siguiente)
        omitir.grid(row=6, column=1, pady=10)

        volver = Button(self.frame_preguntas, text="Volver", font=(self.font, 10), command=self.volver_datos)
        volver.grid(row=6, column=0)

    
    def widgets_puntuacion(self):
        #-------------backgraund and geometry--------------#
        self.frame_puntuacion.config(bg=self.bg)

        #---------------------labels-----------------------#
        label_puntuacion = Label(self.frame_puntuacion, text="Ronda ha terminado", fg="#ffffff", background=self.bg, font=(self.font, 45))
        label_puntuacion.grid(row=0, column=0, pady=10, columnspan=2)

        label_puntuacion = Label(self.frame_puntuacion, text="Puntuación final:", fg="#ffffff", background=self.bg, font=(self.font, 25))
        label_puntuacion.grid(row=1, column=0, pady=10, columnspan=2)

        label_p1 = Label(self.frame_puntuacion, textvariable=self.p1, fg="#ffffff", background=self.bg, font=(self.font, 25))
        label_p1.grid(row=2, column=0, padx=50, pady=10)

        self.label_puntos_1 = Label(self.frame_puntuacion, background=self.bg, fg="#ffffff", font=(self.font, 25))
        self.label_puntos_1.grid(row=3, column=0, pady=10)

        label_p2 = Label(self.frame_puntuacion, textvariable=self.p2, fg="#ffffff", background=self.bg, font=(self.font, 25))
        label_p2.grid(row=2, column=1, padx=50, pady=10)

        self.label_puntos_2 = Label(self.frame_puntuacion, background=self.bg, fg="#ffffff", font=(self.font, 25))
        self.label_puntos_2.grid(row=3, column=1, pady=10)

        #---------------------buttons-----------------------#
        otra_ronda = Button(self.frame_puntuacion, text="Otra ronda", font=(self.font, 12), command=self.otra_ronda)
        otra_ronda.grid(row=4, column=0, pady=10, columnspan=2)

        salir = Button(self.frame_puntuacion, text="Salir", font=(self.font, 12), command=self.salir_ronda)
        salir.grid(row=5, column=0, pady=10, columnspan=2)


    #------------------------functions----------------------------#

    #----------------functions frame inicio-----------------------#

    def comenzar_juego(self):
        #------Se cambia el frame------#
        self.frame_inicio.place_forget()
        self.frame_cargar.place(anchor="c", relx=.5, rely=.5)


    def acerca_de(self):
        messagebox.showinfo("Información adicional", "Desarrollador: Yeison Rojas.\nPrograma para la seleccion del equipo nacional.\nVersión: 1.0\nPrograma bajo licencia MIT")
    
    def salir_programa(self):
        self.root.destroy()

    #----------------functions frame cargar-----------------------#

    def cargar_preguntas(self):
        fichero = filedialog.askopenfilename(title="Abrir", filetypes=(("Ficheros de Excel", "*.xlsx"), ("Todos los ficheros", "*.*")))
        self.direccion.set(fichero)

        self.activar_iniciar()
    
    
    def activar_iniciar(self):
        
        if self.direccion.get() != "":
            self.iniciar.config(state=ACTIVE)
        
        else:
            self.iniciar.config(state=DISABLED)

    
    def cargar_participantes(self):

        #------Se extraen los datos del archivo-----#
        datos = self.leer_archivo()

        self.preguntas = datos[0]
        self.respuestas = datos[1]
        
        #------Se limpian los campos de los participantes-----#
        self.p1.set("")
        self.p2.set("")
        
        #------Se dehabilita el acceso rapido de la tecla Enter-----#
        self.root.bind("<Key-Return>", lambda _: ())
        
        #------Se habilita la lectura de todas las entradas del teclado para validar la cantidad de preguntas-----#
        self.root.bind("<KeyRelease>", lambda _: self.validar_entero())

        #-------Se dehabilita el boton de comenzar ronda-------#
        self.comenzar_ronda.config(state=DISABLED)

        #------Se cambia el frame------#
        self.frame_cargar.place_forget()
        self.frame_datos.place(anchor="c", relx=.5, rely=.5)

    
    def volver_inicio(self):
        self.frame_cargar.place_forget()
        self.frame_inicio.place(anchor="c", relx=.5, rely=.5)
    
    #----------------functions frame participantes-----------------------#

    #------valida si la cantidad de preguntas es un numero entero---------#
    def validar_entero(self):
        codigo = self.c_preguntas.get()

        for i in codigo:
            #-----si no es un numero se borra del entry----#
            if i not in '0123456789':
                self.entry_cantidad.delete(codigo.index(i), codigo.index(i)+1)

        self.activar_comenzar()
    

    def activar_comenzar(self):

        self.comenzar_ronda.config(state=DISABLED)
        
        if self.p1.get() != "" and self.p2.get() != "" and self.c_preguntas.get() != "":
            
            cantidad = int(self.c_preguntas.get())

            if cantidad > 4:
                self.comenzar_ronda.config(state=ACTIVE)
            else:
                self.comenzar_ronda.config(state=DISABLED)

    
    def iniciar_ronda(self):
        
        #------Se dehabilita la lectura de todas las entradas del teclado-------#
        self.root.bind("<KeyRelease>", lambda _: ())

        #------------------acceso directo con el teclado-------------------#

        #----------------Left para participante #1 y Right para participante #2----------------#
        self.root.bind("<Key-Left>", lambda _: self.temporizador(0))
        self.root.bind("<Key-Right>", lambda _: self.temporizador(1))

        #-------Deshabilita los botones de correcto e incorrecto------#
        self.correcto.config(state=DISABLED)
        self.incorrecto.config(state=DISABLED)

        #-------Habilita la modificacion de los campos de texto------#
        self.text_pregunta.config(state="normal")
        self.text_respuesta.config(state="normal")
        
        #-------Borra el contenido anterior------#
        self.text_pregunta.delete("1.0", END)
        self.text_respuesta.delete("1.0", END)

        #-------Valores iniciales de las variables------#
        self.puntos = [0,0]
        self.contador = 0
        self.n_participante = None
        
        #------Se genera el numero aleatoria para extraer una pregunta------#
        self.n = random.randint(0, len(self.preguntas)-1)

        #----Se extrae la pregunta-----#
        p = self.preguntas.pop(self.n)
        r = self.respuestas.pop(self.n)

        #----Se aumenta el contador de preguntas----#
        self.contador +=1
        
        #------Valores necesarios para el temporizador-----#
        self.tmp = False
        self.tiempo.set("")

        #----------Inserta primera pregunta y respuesta-----------#
        self.text_pregunta.insert(INSERT, p)
        self.text_respuesta.insert(INSERT, r)


        #-------Deshabilita la modificacion de los campos de texto------#
        self.text_pregunta.config(state="disabled")
        self.text_respuesta.config(state="disabled")

        #------Se cambia el frame------#
        self.frame_datos.place_forget()
        self.frame_preguntas.place(anchor="c", relx=.5, rely=.5)

    
    def volver_cargar(self):
        #------Se cambia el frame------#
        self.frame_datos.place_forget()
        self.frame_cargar.place(anchor="c", relx=.5, rely=.5)


    #----------------functions frame preguntas-----------------------#
    def refrescar(self):
        if self.tmp == True:
            if self.tiempo.get() > 1:
                self.tiempo.set(self.tiempo.get() - 1)
                
                if self.tmp == False:
                    return
                
                self.root.after(1000, self.refrescar)
            
            else:
                self.tiempo.set("Tiempo acabó")
                #-------Habilita los botones de responder y los cambia a su color original------#
                self.responder_1.config(state=ACTIVE, background='#f1eff0')
                self.responder_2.config(state=ACTIVE, background='#f1eff0')

                #----------------Left para participante #1 y Right para participante #2----------------#
                self.root.bind("<Key-Left>", lambda _: self.temporizador(0))
                self.root.bind("<Key-Right>", lambda _: self.temporizador(1))

                #-------------Activa tecla Up y Down para marcar respuesta correcta o incorrecta---------------#
                self.root.bind("<Key-Up>", lambda _: self.r_correcta())
                self.root.bind("<Key-Down>", lambda _: self.r_incorrecta())  
        
        else:
            #-------Habilita los botones de responder y los cambia a su color original------#
            self.responder_1.config(state=ACTIVE, background='#f1eff0')
            self.responder_2.config(state=ACTIVE, background='#f1eff0')

            #----------Desactiva tecla Up para respuesta correcta-----------#
            self.root.bind("<Key-Up>", lambda _: ())

            #---------Desactiva tecla Down para respuesta incorrecta--------#
            self.root.bind("<Key-Down>", lambda _: ())
        
    
    def temporizador(self, n):

        self.n_participante = n
        
        #-------Desactiva el boton responder y acceso rapido desde el teclado------#
        self.responder_1.config(state=DISABLED)
        self.responder_2.config(state=DISABLED)

        self.root.bind("<Key-Left>", lambda _: ())
        self.root.bind("<Key-Right>", lambda _: ())
        

        #-------Cambia a su color del boton presionado------#
        if self.n_participante == 0:
            self.responder_1.config(background='#54FA9B')
        
        elif self.n_participante == 1:
            self.responder_2.config(background='#54FA9B')


        #-------Activa los botones correcto e incorrecto------#
        self.correcto.config(state=ACTIVE)
        self.incorrecto.config(state=ACTIVE)

        #-------------Activa tecla Up para marcar respuesta correcta---------------#
        self.root.bind("<Key-Up>", lambda _: self.r_correcta())

        #-------------Activa tecla Down para marcar respuesta incorrecta---------------#
        self.root.bind("<Key-Down>", lambda _: self.r_incorrecta())
        
        self.tmp = True
        self.tiempo.set("6")
        self.refrescar()

    def r_correcta(self):
        self.puntos[self.n_participante] += 5
        self.contador +=1
        self.siguiente()
    
    def r_incorrecta(self):
        self.puntos[self.n_participante] -= 5
        self.contador +=1
        self.siguiente()

    def siguiente(self):

        self.tmp = False
        self.tiempo.set("")

        self.n_participante = None

        #------Habilita el acceso rapido desde el teclado---------#
        self.root.bind("<Key-Left>", lambda _: self.temporizador(0))
        self.root.bind("<Key-Right>", lambda _: self.temporizador(1))

        #-------Deshabilita los botones de correcto e incorrecto------#
        self.correcto.config(state=DISABLED)
        self.incorrecto.config(state=DISABLED)
        
        #-------Habilita la modificacion de los campos de texto------#
        self.text_pregunta.config(state="normal")
        self.text_respuesta.config(state="normal")

        #-------Borra el contenido anterior------#
        self.text_pregunta.delete("1.0", END)
        self.text_respuesta.delete("1.0", END)

        cantidad = int(self.c_preguntas.get())

        if len(self.preguntas) and self.contador <= cantidad:

            self.n = random.randint(0, len(self.preguntas)-1)

            p = self.preguntas.pop(self.n)
            r = self.respuestas.pop(self.n)

            #---------Inserta otra pregunta----------#
            self.text_pregunta.insert(INSERT, p)
            self.text_respuesta.insert(INSERT, r)

            #-------Deshabilita la modificacion de los campos de texto------#
            self.text_pregunta.config(state="disabled")
            self.text_respuesta.config(state="disabled")

        else:

            #----------------Desactiva el acceso rapido desde el teclado----------------#
            self.root.bind("<Key-Return>", lambda _: ())

            self.root.bind("<Key-Left>", lambda _: ())
            self.root.bind("<Key-Right>", lambda _: ())

            self.root.bind("<Key-Up>", lambda _: ())
            self.root.bind("<Key-Down>", lambda _: ())
            
            self.label_puntos_1.config(text=self.puntos[0])
            self.label_puntos_2.config(text=self.puntos[1])

            #------Se cambia el frame------#
            self.frame_preguntas.place_forget()
            self.frame_puntuacion.place(anchor="c", relx=.5, rely=.5)


    def leer_archivo(self):
        
        excel_dataframe = openpyxl.load_workbook(self.direccion.get())

        dataframe = excel_dataframe.active

        data = []

        for row in range(0, dataframe.max_row):

            for col in dataframe.iter_cols(0, dataframe.max_column):

                data.append(col[row].value)


        preguntas = []
        respuestas = []

        x = 0

        while x < dataframe.max_row:
            preguntas.append(data[x])
            respuestas.append(data[x+1])
            x += 3

        return [preguntas, respuestas]

    
    def volver_datos(self):

        #----Se habilita la lectura de todas las entradas del teclado----#
        self.root.bind("<KeyRelease>", lambda _: self.validar_entero())

        self.p1.set("")
        self.p2.set("")

        #------Se cambia el frame------#
        self.frame_preguntas.place_forget()
        self.frame_datos.place(anchor="c", relx=.5, rely=.5)
    
    
    #----------------functions frame puntuacion-----------------------#

    def otra_ronda(self):

        if len(self.preguntas) == 0:
            self.salir_ronda()
            
        else:
            self.frame_puntuacion.place_forget()
            self.frame_datos.place(anchor="c", relx=.5, rely=.5)
        
            self.p1.set("")
            self.p2.set("")
            self.root.bind("<KeyRelease>", lambda _: self.validar_entero())
            self.comenzar_ronda.config(state=DISABLED)

    
    def salir_ronda(self):
        self.frame_puntuacion.place_forget()
        self.frame_inicio.place(anchor="c", relx=.5, rely=.5)
        self.p1.set("")
        self.p2.set("")


app = programa_spe()
app.root.mainloop()
